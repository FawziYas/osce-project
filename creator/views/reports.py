"""
Reports views – scoresheets, report index, and XLSX exports.
"""
from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseForbidden
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Sum, F
from collections import defaultdict

from core.models import (
    Course, Exam, ExamSession, SessionStudent, Station, ChecklistItem,
    StationScore, ItemScore, ILO,
)
from core.utils.roles import scope_queryset, check_session_department


@login_required
def reports_index(request):
    """Reports dashboard – select session and generate reports. Dept-scoped."""
    # Base filter
    sessions_qs = ExamSession.objects.filter(
        exam__is_deleted=False,
    ).select_related('exam', 'exam__course')
    
    # Department scoping
    sessions_qs = scope_queryset(request.user, sessions_qs, dept_field='exam__course__department')
    
    # Permission-based filtering
    if not request.user.is_superuser:
        sessions_qs = sessions_qs.filter(status='completed')
    
    sessions = sessions_qs.order_by('-session_date')

    selected_session = None
    session_id = request.GET.get('session_id')
    if session_id:
        selected_session = ExamSession.objects.filter(pk=session_id).first()

    return render(request, 'creator/reports/index.html', {
        'sessions': sessions,
        'selected_session': selected_session,
    })


@login_required
def reports_scoresheets(request, session_id):
    """Print-ready score sheets for a session — dept-scoped."""
    session = get_object_or_404(ExamSession, pk=session_id)
    
    # Department scoping
    if not check_session_department(request.user, session):
        return HttpResponseForbidden('You do not have access to this session.')
    
    # Access control: non-superusers can only view completed sessions
    if not request.user.is_superuser and session.status != 'completed':
        return HttpResponseForbidden(
            "You do not have permission to view this session. "
            "Only superusers can view non-completed sessions."
        )
    
    # Search filter
    search_query = request.GET.get('search', '').strip()
    print_all = request.GET.get('print_all') == '1'

    students_qs = SessionStudent.objects.filter(session=session)
    if search_query:
        students_qs = students_qs.filter(full_name__icontains=search_query) | SessionStudent.objects.filter(
            session=session, student_number__icontains=search_query
        )
    students_qs = students_qs.order_by('full_name').distinct()

    # Paginate: 10 per page normally; all students in print_all mode
    per_page = students_qs.count() if print_all else 10
    page_num = 1 if print_all else request.GET.get('page', 1)
    paginator = Paginator(students_qs, per_page=max(per_page, 1))

    try:
        students_page = paginator.page(page_num)
    except (PageNotAnInteger, EmptyPage):
        students_page = paginator.page(1)

    students = students_page.object_list

    student_data = []
    for student in students:
        if student.path_id:
            student_stations = Station.objects.filter(
                path_id=student.path_id, active=True,
            ).order_by('station_number')
        else:
            student_stations = Station.objects.none()

        # Only count submitted scores, ignore abandoned/in-progress evaluations
        scores_qs = StationScore.objects.filter(
            session_student=student,
            status='submitted'
        ).exclude(
            total_score__isnull=True
        ).select_related('examiner')
        # Map all scores by station (not just one per station)
        score_map_all = {}
        for score in scores_qs:
            if score.station_id not in score_map_all:
                score_map_all[score.station_id] = []
            score_map_all[score.station_id].append(score)

        max_possible = 0
        for st in student_stations:
            st_max = ChecklistItem.objects.filter(station=st).aggregate(
                total=Sum('points')
            )['total'] or 0
            max_possible += st_max

        # Calculate total using average for multiple examiners per station
        total_score = 0
        for station_id, scores_list in score_map_all.items():
            if scores_list:
                # Use average if multiple scores, otherwise use single score
                avg_score = sum(s.total_score or 0 for s in scores_list) / len(scores_list)
                total_score += avg_score
        
        percentage = (total_score / max_possible * 100) if max_possible > 0 else 0
        
        # Build comments with examiner names
        comments_with_examiner = []
        for score in scores_qs:
            if score.comments and score.comments.strip():
                examiner_name = score.examiner.full_name if score.examiner else "Unknown Examiner"
                comments_with_examiner.append({
                    'station': Station.objects.filter(pk=score.station_id).first(),
                    'examiner_name': examiner_name,
                    'comment_text': score.comments,
                    'formatted': f"{examiner_name}:\n{score.comments}"
                })
        
        student_data.append({
            'student': student,
            'stations': student_stations,
            'scores': score_map_all,
            'total_score': round(total_score, 2),
            'max_score': max_possible,
            'percentage_display': round(percentage, 2),
            'passed': percentage >= 60,
            'comments_with_examiner': comments_with_examiner,
        })

    return render(request, 'creator/reports/scoresheets.html', {
        'session': session,
        'students': student_data,
        'page_obj': students_page,
        'paginator': paginator,
        'search_query': search_query,
        'print_all': print_all,
    })


@login_required
def reports_student_scoresheet(request, student_id):
    """Print-ready score sheet for a single student — dept-scoped."""
    student = get_object_or_404(SessionStudent, pk=student_id)
    session = get_object_or_404(ExamSession, pk=student.session_id)
    
    # Department scoping
    if not check_session_department(request.user, session):
        return HttpResponseForbidden('You do not have access to this session.')
    
    # Access control: non-superusers can only view completed sessions
    if not request.user.is_superuser and session.status != 'completed':
        return HttpResponseForbidden(
            "You do not have permission to view this session. "
            "Only superusers can view non-completed sessions."
        )

    if student.path_id:
        student_stations = Station.objects.filter(
            path_id=student.path_id, active=True,
        ).order_by('station_number')
    else:
        student_stations = Station.objects.none()

    # Only count submitted scores, ignore abandoned/in-progress evaluations
    scores_qs = StationScore.objects.filter(
        session_student=student,
        status='submitted'
    ).exclude(
        total_score__isnull=True
    ).select_related('examiner')
    # Map all scores by station (not just one per station)
    score_map_all = {}
    for score in scores_qs:
        if score.station_id not in score_map_all:
            score_map_all[score.station_id] = []
        score_map_all[score.station_id].append(score)

    max_possible = 0
    for st in student_stations:
        st_max = ChecklistItem.objects.filter(station=st).aggregate(
            total=Sum('points')
        )['total'] or 0
        max_possible += st_max

    # Calculate total using average for multiple examiners per station
    _total = 0
    for station_id, scores_list in score_map_all.items():
        if scores_list:
            # Use average if multiple scores, otherwise use single score
            avg_score = sum(s.total_score or 0 for s in scores_list) / len(scores_list)
            _total += avg_score
    
    _pct = (_total / max_possible * 100) if max_possible > 0 else 0
    
    # Build comments with examiner names
    comments_with_examiner = []
    for score in scores_qs:
        if score.comments and score.comments.strip():
            examiner_name = score.examiner.full_name if score.examiner else "Unknown Examiner"
            comments_with_examiner.append({
                'station': Station.objects.filter(pk=score.station_id).first(),
                'examiner_name': examiner_name,
                'comment_text': score.comments,
                'formatted': f"{examiner_name}:\n{score.comments}"
            })
    
    student_data = [{
        'student': student,
        'stations': student_stations,
        'scores': score_map_all,
        'total_score': round(_total, 2),
        'max_score': max_possible,
        'percentage_display': round(_pct, 2),
        'passed': _pct >= 60,
        'comments_with_examiner': comments_with_examiner,
    }]

    return render(request, 'creator/reports/scoresheets.html', {
        'session': session,
        'students': student_data,
        'single_student': True,
    })


# ── XLSX Export: ILO Scores per Student ─────────────────────────────────

@login_required
def export_ilo_scores_xlsx(request, session_id):
    """
    Generate an XLSX workbook for a session. Dept-scoped.
    Sheet layout — rows = students, columns = ILOs.
    Each cell = average of examiner scores for that student / ILO.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    session = get_object_or_404(
        ExamSession.objects.select_related('exam', 'exam__course'),
        pk=session_id,
    )

    # Department scoping
    if not check_session_department(request.user, session):
        return HttpResponseForbidden('You do not have access to this session.')

    # Access control
    if not request.user.is_superuser and session.status != 'completed':
        return HttpResponseForbidden("Only superusers can export non-completed sessions.")

    exam = session.exam
    course = exam.course

    # ── Gather ILOs for this course ─────────────────────────────────────
    ilos = list(ILO.objects.filter(course=course).order_by('number'))
    if not ilos:
        return HttpResponse("No ILOs defined for this course.", status=400)

    ilo_ids = [ilo.id for ilo in ilos]

    # ── Gather students ─────────────────────────────────────────────────
    students = list(
        SessionStudent.objects.filter(session=session)
        .order_by('full_name')
    )
    student_ids = [s.id for s in students]

    # ── Fetch all submitted item-level scores in one query ──────────────
    # Join: ItemScore → StationScore → SessionStudent
    #        ItemScore → ChecklistItem (for ilo_id, points)
    rows = (
        ItemScore.objects
        .filter(
            station_score__session_student_id__in=student_ids,
            station_score__status='submitted',
            checklist_item__ilo_id__in=ilo_ids,
        )
        .values(
            'station_score__session_student_id',
            'station_score__station_id',
            'station_score__examiner_id',
            'checklist_item__ilo_id',
        )
        .annotate(
            earned=Sum('score'),
            possible=Sum('max_points'),
        )
    )

    # Structure:  student_id → ilo_id → list of (earned, possible) per examiner-station combo
    # When multiple examiners score the same station we average their totals.
    # raw_data[student_id][ilo_id][(station_id, examiner_id)] = (earned, possible)
    raw_data = defaultdict(lambda: defaultdict(dict))
    for r in rows:
        sid = r['station_score__session_student_id']
        ilo_id = r['checklist_item__ilo_id']
        key = (r['station_score__station_id'], r['station_score__examiner_id'])
        raw_data[sid][ilo_id][key] = (r['earned'] or 0, r['possible'] or 0)

    # Collapse to averages per station, then sum across stations.
    # Result: scores[student_id][ilo_id] = {'earned': float, 'possible': float}
    scores = {}
    for sid in student_ids:
        scores[sid] = {}
        for ilo in ilos:
            examiner_station_data = raw_data[sid].get(ilo.id, {})
            if not examiner_station_data:
                scores[sid][ilo.id] = {'earned': 0, 'possible': 0}
                continue

            # Group by station, average examiners within each station
            station_groups = defaultdict(list)
            station_possible = {}
            for (station_id, _examiner_id), (earned, possible) in examiner_station_data.items():
                station_groups[station_id].append(earned)
                station_possible[station_id] = possible  # same max for same station

            total_earned = 0
            total_possible = 0
            for station_id, earned_list in station_groups.items():
                total_earned += sum(earned_list) / len(earned_list)  # average across examiners
                total_possible += station_possible[station_id]

            scores[sid][ilo.id] = {
                'earned': round(total_earned, 2),
                'possible': round(total_possible, 2),
            }

    # ── Build workbook ──────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'ILO Scores'

    # Styles
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1A1A2E', end_color='1A1A2E', fill_type='solid')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center')

    # ── Row 1: Title ────────────────────────────────────────────────────
    total_cols = 3 + len(ilos) + 1  # fixed + 1 col per ILO + Total
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_cell = ws.cell(row=1, column=1,
                         value=f'{exam.name} — {session.name} — ILO Score Report')
    title_cell.font = Font(bold=True, size=14, color='1A1A2E')
    title_cell.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 30

    # ── Row 2: Course info ──────────────────────────────────────────────
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    info_cell = ws.cell(row=2, column=1,
                        value=f'Course: {course.code} — {course.name}  |  Students: {len(students)}')
    info_cell.font = Font(size=10, color='666666')
    ws.row_dimensions[2].height = 20

    # ── Row 3: Empty spacer ─────────────────────────────────────────────
    start_row = 4

    # ── Row 4: Header ───────────────────────────────────────────────────
    fixed_headers = ['#', 'Student Number', 'Student Name']
    for ci, hdr in enumerate(fixed_headers, 1):
        cell = ws.cell(row=start_row, column=ci, value=hdr)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    # One column per ILO
    col = len(fixed_headers) + 1
    for ilo in ilos:
        cell = ws.cell(row=start_row, column=col, value=f'ILO #{ilo.number}')
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        col += 1

    # Total column
    total_col = col
    total_hdr = ws.cell(row=start_row, column=total_col, value='TOTAL')
    total_hdr.font = header_font
    total_hdr.fill = PatternFill(start_color='198754', end_color='198754', fill_type='solid')
    total_hdr.alignment = center
    total_hdr.border = border

    # ── Data rows ───────────────────────────────────────────────────────
    data_start = start_row + 1
    even_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')

    for idx, student in enumerate(students):
        row_num = data_start + idx
        row_fill = even_fill if idx % 2 == 0 else None

        # Fixed columns
        ws.cell(row=row_num, column=1, value=idx + 1).alignment = center
        ws.cell(row=row_num, column=2, value=student.student_number).alignment = left_align
        ws.cell(row=row_num, column=3, value=student.full_name).alignment = left_align

        for ci in range(1, 4):
            ws.cell(row=row_num, column=ci).border = border
            if row_fill:
                ws.cell(row=row_num, column=ci).fill = row_fill

        # ILO score columns
        col = len(fixed_headers) + 1
        grand_earned = 0

        for ilo in ilos:
            s = scores.get(student.id, {}).get(ilo.id, {'earned': 0, 'possible': 0})
            earned = s['earned']
            grand_earned += earned

            earned_cell = ws.cell(row=row_num, column=col, value=earned if earned else '')
            earned_cell.alignment = center
            earned_cell.border = border
            if row_fill:
                earned_cell.fill = row_fill
            col += 1

        # Total
        total_cell = ws.cell(row=row_num, column=total_col, value=round(grand_earned, 2))
        total_cell.font = Font(bold=True)
        total_cell.alignment = center
        total_cell.border = border
        if row_fill:
            total_cell.fill = row_fill

    # ── ILO description row at the bottom ───────────────────────────────
    desc_row = data_start + len(students) + 1
    ws.cell(row=desc_row, column=1, value='ILO Descriptions:').font = Font(bold=True, size=10)
    for i, ilo in enumerate(ilos):
        ws.cell(row=desc_row + 1 + i, column=1,
                value=f'ILO #{ilo.number}: {ilo.description} (Max: {ilo.osce_marks})')
        ws.cell(row=desc_row + 1 + i, column=1).font = Font(size=9, color='555555')

    # ── Column widths ───────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 28
    for ci in range(len(fixed_headers) + 1, total_col + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 10

    # ── Freeze panes ────────────────────────────────────────────────────
    ws.freeze_panes = f'D{data_start}'

    # ── Response ────────────────────────────────────────────────────────
    filename = f'ILO_Scores_{exam.name}_{session.name}.xlsx'.replace(' ', '_')
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
