"""
Examiner management views – list, CRUD, unassign, bulk upload/download template.
"""
from datetime import datetime
from io import BytesIO
from django.utils import timezone
from django.conf import settings

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse

from core.models import Examiner, ExaminerAssignment, ExamSession, Department
from core.utils.roles import scope_queryset
from core.utils.cache_utils import (
    get_departments, invalidate_examiner_list, invalidate_departments,
    EXAMINER_LIST_KEY, EXAMINER_LIST_TTL,
    EXAMINER_STATS_KEY, EXAMINER_STATS_TTL,
)


@login_required
def examiner_list(request):
    """List all examiners (active only). Deleted shown separately for admin/superuser."""
    from django.core.cache import cache
    from core.utils.roles import is_global

    # Department-scoped query (cache only for global users)
    if is_global(request.user):
        examiners = cache.get(EXAMINER_LIST_KEY)
        if examiners is None:
            examiners = list(
                Examiner.objects.filter(role='examiner', is_deleted=False).order_by('full_name')
            )
            cache.set(EXAMINER_LIST_KEY, examiners, EXAMINER_LIST_TTL)
    else:
        examiners = list(scope_queryset(
            request.user,
            Examiner.objects.filter(role='examiner', is_deleted=False),
            dept_field='department',
        ).order_by('full_name'))

    today = datetime.now().date()

    can_see_deleted = request.user.is_superuser or getattr(request.user, 'role', None) == 'admin'
    deleted_examiners = Examiner.objects.filter(role='examiner', is_deleted=True).order_by('full_name') if can_see_deleted else []

    stats = cache.get(EXAMINER_STATS_KEY)
    if stats is None:
        stats = {
            'total': Examiner.objects.filter(role='examiner', is_deleted=False).count(),
            'active': Examiner.objects.filter(role='examiner', is_active=True, is_deleted=False).count(),
            'assigned_today': ExaminerAssignment.objects.filter(
                session__session_date=today,
                examiner__role='examiner',
                examiner__is_deleted=False,
            ).values('examiner_id').distinct().count(),
            'total_assignments': ExaminerAssignment.objects.filter(
                examiner__role='examiner',
                examiner__is_deleted=False,
            ).count(),
        }
        cache.set(EXAMINER_STATS_KEY, stats, EXAMINER_STATS_TTL)

    return render(request, 'creator/examiners/list.html', {
        'examiners': examiners,
        'stats': stats,
        'deleted_examiners': deleted_examiners,
        'can_see_deleted': can_see_deleted,
    })


@login_required
def examiner_create(request):
    """Create a new examiner."""
    if request.method == 'POST':
        username = request.POST['username'].strip().lower()
        email = request.POST.get('email', '').strip().lower()

        if Examiner.objects.filter(username=username).exists():
            messages.error(request, f'Username "{username}" already exists.')
            default_password = getattr(settings, 'DEFAULT_USER_PASSWORD', '12345678F')
            departments = get_departments()
            return render(request, 'creator/examiners/form.html', {
                'examiner': None,
                'default_password': default_password,
                'departments': departments,
            })

        if email and Examiner.objects.filter(email=email).exists():
            messages.error(request, f'Email "{email}" already exists.')
            default_password = getattr(settings, 'DEFAULT_USER_PASSWORD', '12345678F')
            departments = get_departments()
            return render(request, 'creator/examiners/form.html', {
                'examiner': None,
                'default_password': default_password,
                'departments': departments,
            })

        examiner = Examiner(
            username=username,
            email=email,
            full_name=request.POST['full_name'].strip(),
            title=request.POST.get('title', '').strip() or '',
            is_active='is_active' in request.POST,
        )
        # Set department FK by looking up the Department by name
        dept_name = request.POST.get('department', '').strip()
        if dept_name:
            try:
                examiner.department = Department.objects.get(name=dept_name)
            except Department.DoesNotExist:
                pass
        # Password is set automatically to DEFAULT_USER_PASSWORD by the
        # post_save signal.  User will be forced to change it on first login.
        examiner.save()
        invalidate_examiner_list()

        messages.success(request, f'Examiner "{examiner.display_name}" created successfully.')
        return redirect('creator:examiner_list')

    default_password = getattr(settings, 'DEFAULT_USER_PASSWORD', '12345678F')
    departments = get_departments()
    return render(request, 'creator/examiners/form.html', {
        'examiner': None,
        'default_password': default_password,
        'departments': departments,
    })


@login_required
def examiner_detail(request, examiner_id):
    """View examiner details and assignments."""
    examiner = get_object_or_404(
        scope_queryset(request.user, Examiner.objects.all(), dept_field='department'),
        pk=examiner_id,
    )
    assignments = ExaminerAssignment.objects.filter(
        examiner=examiner
    ).select_related('session', 'station')

    return render(request, 'creator/examiners/detail.html', {
        'examiner': examiner,
        'assignments': assignments,
    })


@login_required
def examiner_edit(request, examiner_id):
    """Edit an examiner."""
    examiner = get_object_or_404(
        scope_queryset(request.user, Examiner.objects.all(), dept_field='department'),
        pk=examiner_id,
    )

    if request.method == 'POST':
        email = request.POST.get('email', '').strip().lower()
        if email and Examiner.objects.filter(email=email).exclude(pk=examiner_id).exists():
            messages.error(request, f'Email "{email}" is already used by another examiner.')
            departments = get_departments()
            return render(request, 'creator/examiners/form.html', {
                'examiner': examiner,
                'departments': departments,
            })

        examiner.email = email
        examiner.full_name = request.POST['full_name'].strip()
        examiner.title = request.POST.get('title', '').strip() or ''
        # Set department FK by looking up the Department by name
        dept_name = request.POST.get('department', '').strip()
        if dept_name:
            try:
                examiner.department = Department.objects.get(name=dept_name)
            except Department.DoesNotExist:
                examiner.department = None
        else:
            examiner.department = None
        examiner.is_active = 'is_active' in request.POST

        examiner.save()
        invalidate_examiner_list()
        messages.success(request, f'Examiner "{examiner.display_name}" updated.')
        return redirect('creator:examiner_detail', examiner_id=examiner_id)

    departments = get_departments()
    return render(request, 'creator/examiners/form.html', {
        'examiner': examiner,
        'departments': departments,
    })


@login_required
def examiner_delete(request, examiner_id):
    """Soft-delete an examiner. Admin/superuser only."""
    if not (request.user.is_superuser or getattr(request.user, 'role', None) == 'admin'):
        messages.error(request, 'Only admins and superusers can delete examiners.')
        return redirect('creator:examiner_list')

    if request.method != 'POST':
        messages.error(request, 'Invalid request method.')
        return redirect('creator:examiner_list')

    examiner = get_object_or_404(Examiner, pk=examiner_id, is_deleted=False)
    name = examiner.display_name
    examiner.is_deleted = True
    examiner.is_active = False
    examiner.deleted_at = timezone.now()
    examiner.save()
    invalidate_examiner_list()

    messages.success(request, f'Examiner "{name}" has been deleted. You can restore them from the deleted list.')
    return redirect('creator:examiner_list')


@login_required
def examiner_restore(request, examiner_id):
    """Restore a soft-deleted examiner. Admin/superuser only."""
    if not (request.user.is_superuser or getattr(request.user, 'role', None) == 'admin'):
        messages.error(request, 'Only admins and superusers can restore examiners.')
        return redirect('creator:examiner_list')

    if request.method != 'POST':
        return redirect('creator:examiner_list')

    examiner = get_object_or_404(Examiner, pk=examiner_id, is_deleted=True)
    examiner.is_deleted = False
    examiner.is_active = True
    examiner.deleted_at = None
    examiner.save()
    invalidate_examiner_list()

    messages.success(request, f'Examiner "{examiner.display_name}" has been restored.')
    return redirect('creator:examiner_list')


@login_required
def examiner_permanent_delete(request, examiner_id):
    """Permanently delete an examiner. Superuser only.
    
    Station assignments are removed, but all scoring records are preserved
    with examiner shown as 'Deleted Examiner' in reports.
    """
    if not request.user.is_superuser:
        messages.error(request, 'Only superusers can permanently delete examiners.')
        return redirect('creator:examiner_list')

    if request.method != 'POST':
        return redirect('creator:examiner_list')

    examiner = get_object_or_404(Examiner, pk=examiner_id, is_deleted=True)
    name = examiner.display_name
    # StationScore records are preserved (SET_NULL on examiner FK)
    # ExaminerAssignment records are removed via CASCADE
    examiner.delete()
    invalidate_examiner_list()

    messages.success(request, f'Examiner "{name}" permanently deleted. Their scoring records have been preserved.')
    return redirect('creator:examiner_list')


@login_required
def examiner_unassign(request, assignment_id):
    """Remove an examiner assignment."""
    assignment = get_object_or_404(ExaminerAssignment, pk=assignment_id)
    examiner_id = assignment.examiner_id

    session = ExamSession.objects.filter(pk=assignment.session_id).first()
    if session and session.actual_start and not request.user.is_superuser:
        messages.error(request, 'Cannot remove examiner assignments after session has been activated.')
        return redirect('creator:session_detail', session_id=str(session.id))

    session_id = assignment.session_id
    assignment.delete()
    messages.success(request, 'Assignment removed.')

    # Redirect back to where the user came from
    next_url = request.POST.get('next') or request.GET.get('next', '')
    if next_url:
        from django.utils.http import url_has_allowed_host_and_scheme
        if url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}):
            return redirect(next_url)
    # If coming from session detail, go back there
    referer = request.META.get('HTTP_REFERER', '')
    if session_id and 'session' in referer:
        return redirect('creator:session_detail', session_id=str(session_id))
    return redirect('creator:examiner_detail', examiner_id=examiner_id)


# ---------------------------------------------------------------------------
# Bulk operations
# ---------------------------------------------------------------------------

@login_required
def examiner_download_template(request):
    """Download an XLSX template for bulk examiner upload with department dropdowns."""
    if not request.user.is_superuser:
        messages.error(request, 'You do not have permission to download templates.')
        return redirect('creator:examiner_list')
    
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.worksheet.datavalidation import DataValidation
    from core.models import Department

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Examiners Template'

    headers = ['title', 'full_name', 'username', 'email', 'department']
    arabic_hints = [
        'اللقب (مثلاً Dr.)', 'الاسم الكامل', 'اسم المستخدم',
        'البريد الإلكتروني', 'القسم',
    ]

    header_fill = PatternFill(start_color='CFE2F3', end_color='CFE2F3', fill_type='solid')
    header_font = Font(bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

        hint = ws.cell(row=2, column=col, value=arabic_hints[col - 1])
        hint.font = Font(italic=True, color='666666')
        hint.alignment = Alignment(horizontal='right')

    sample = ['Dr.', 'Ahmed Mansour', 'ahmed_m', 'ahmed@example.com', 'General Surgery']
    for col, val in enumerate(sample, 1):
        ws.cell(row=3, column=col, value=val)

    # Add department dropdown validation
    departments = list(Department.objects.values_list('name', flat=True).order_by('name'))
    if departments:
        # Create a hidden sheet to store department list
        dv_sheet = wb.create_sheet('Departments')
        for idx, dept in enumerate(departments, 1):
            dv_sheet.cell(row=idx, column=1, value=dept)
        
        # Create data validation with dropdown for department column (rows 4-1000)
        dv = DataValidation(
            type='list',
            formula1=f'=Departments!$A$1:$A${len(departments)}',
            allow_blank=True
        )
        dv.error = 'Please select a valid department from the list'
        dv.errorTitle = 'Invalid Department'
        ws.add_data_validation(dv)
        dv.add(f'E4:E1000')  # Apply to department column, rows 4-1000

    for i in range(1, 6):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 20

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'examiner_upload_template_{datetime.now():%Y%m%d}.xlsx'
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response


@login_required
def examiner_bulk_upload(request):
    """Handle XLSX upload for bulk examiner creation with department validation."""
    if not request.user.is_superuser:
        messages.error(request, 'You do not have permission to upload examiners.')
        return redirect('creator:examiner_list')
    
    import openpyxl
    from core.models import Department

    if request.method != 'POST' or 'file' not in request.FILES:
        messages.error(request, 'No file uploaded')
        return redirect('creator:examiner_list')

    f = request.FILES['file']
    if not f.name.endswith('.xlsx'):
        messages.error(request, 'Please upload an .xlsx file')
        return redirect('creator:examiner_list')

    # File size validation (5 MB max)
    max_size = 5 * 1024 * 1024
    if f.size > max_size:
        messages.error(request, 'File too large. Maximum size is 5 MB.')
        return redirect('creator:examiner_list')

    try:
        # Get valid departments
        valid_departments = set(Department.objects.values_list('name', flat=True))
        
        wb = openpyxl.load_workbook(f)
        ws = wb.active

        headers = [str(c.value).strip().lower() for c in ws[1] if c.value]
        required = ['full_name', 'username']
        for field in required:
            if field not in headers:
                messages.error(request, f'Missing required column: {field}')
                return redirect('creator:examiner_list')

        idx = {h: i for i, h in enumerate(headers)}
        success_count = 0
        errors_list = []

        for row_num, row in enumerate(ws.iter_rows(min_row=3, values_only=True), 3):
            if not any(row):
                continue
            try:
                username = str(row[idx['username']]).strip().lower()
                email = str(row[idx['email']]).strip().lower() if 'email' in idx and row[idx['email']] else ''
                full_name = str(row[idx['full_name']]).strip()
                title = str(row[idx.get('title', -1)]).strip() if 'title' in idx and row[idx['title']] else ''
                department = str(row[idx.get('department', -1)]).strip() if 'department' in idx and row[idx['department']] else ''

                if not all([username, full_name]):
                    errors_list.append(f'Row {row_num}: Missing required data')
                    continue

                # Validate department
                if department and department not in valid_departments:
                    errors_list.append(
                        f"Row {row_num}: Department '{department}' does not exist."
                    )
                    continue

                if Examiner.objects.filter(username=username).exists():
                    errors_list.append(f"Row {row_num}: Username '{username}' already exists")
                    continue
                if email and Examiner.objects.filter(email=email).exists():
                    errors_list.append(f"Row {row_num}: Email '{email}' already exists")
                    continue

                # Look up department FK
                dept_obj = None
                if department:
                    try:
                        dept_obj = Department.objects.get(name=department)
                    except Department.DoesNotExist:
                        pass
                new_examiner = Examiner(
                    username=username,
                    email=email,
                    full_name=full_name,
                    title=title,
                    department=dept_obj,
                    is_active=True,
                )
                # Password is set automatically by the post_save signal
                new_examiner.save()
                success_count += 1
            except Exception as e:
                errors_list.append(f'Row {row_num}: {e}')

        if success_count:
            invalidate_examiner_list()
            messages.success(request, f'Successfully imported {success_count} examiners.')
        if errors_list:
            messages.warning(request, f'Failed to import {len(errors_list)} rows.')
            for err in errors_list[:5]:
                messages.error(request, err)
            if len(errors_list) > 5:
                messages.error(request, f'...and {len(errors_list) - 5} more errors.')

    except Exception as e:
        messages.error(request, f'Error processing file: {e}')

    return redirect('creator:examiner_list')


@login_required
def bulk_import_status(request):
    """
    Poll the status of an async bulk examiner import task.
    GET /examiners/bulk-import-status/?task_id=<uuid>
    Returns: {status: 'running'|'done'|'error'|'pending', progress: int, ...}
    """
    from django.core.cache import cache
    from django.http import JsonResponse

    if not request.user.is_superuser:
        return JsonResponse({'status': 'error', 'message': 'Forbidden'}, status=403)

    task_id = request.GET.get('task_id', '')
    if not task_id:
        return JsonResponse({'status': 'error', 'message': 'Missing task_id'}, status=400)

    result = cache.get(f'osce:bulk_import:{task_id}')
    if result is None:
        return JsonResponse({'status': 'pending'})
    return JsonResponse(result)


# ── Coordinator management (admin/superuser only) ─────────────────────────────


@login_required
def coordinator_list(request):
    """List all coordinators with search. Admin/superuser only."""
    if not (request.user.is_superuser or getattr(request.user, 'role', None) == 'admin'):
        messages.error(request, 'You do not have permission to manage coordinators.')
        return redirect('creator:dashboard')

    search_query = request.GET.get('search', '').strip()
    dept_filter = request.GET.get('department', '').strip()

    coordinators_qs = Examiner.objects.filter(role='coordinator').select_related(
        'department'
    ).order_by('department__name', 'full_name')

    if search_query:
        from django.db.models import Q
        coordinators_qs = coordinators_qs.filter(
            Q(full_name__icontains=search_query) |
            Q(username__icontains=search_query) |
            Q(email__icontains=search_query)
        )
    if dept_filter:
        coordinators_qs = coordinators_qs.filter(department_id=dept_filter)

    departments = Department.objects.order_by('name')

    return render(request, 'creator/coordinators/list.html', {
        'coordinators': coordinators_qs,
        'search_query': search_query,
        'departments': departments,
        'dept_filter': dept_filter,
    })


@login_required
def coordinator_create(request):
    """Create a new coordinator with department and position. Admin/superuser only."""
    if not (request.user.is_superuser or getattr(request.user, 'role', None) == 'admin'):
        messages.error(request, 'You do not have permission to add coordinators.')
        return redirect('creator:dashboard')

    default_password = getattr(settings, 'DEFAULT_USER_PASSWORD', '12345678F')
    departments = Department.objects.order_by('name')

    if request.method == 'POST':
        username = request.POST.get('username', '').strip().lower()
        email = request.POST.get('email', '').strip().lower()
        full_name = request.POST.get('full_name', '').strip()
        dept_id = request.POST.get('department', '').strip()
        position = request.POST.get('coordinator_position', '').strip()

        errors = []
        if not username:
            errors.append('Username is required.')
        if not full_name:
            errors.append('Full name is required.')
        if not dept_id:
            errors.append('Department is required.')
        if position not in ('head', 'rta', 'organizer'):
            errors.append('Position must be Head, RTA, or Organizer.')
        if Examiner.objects.filter(username=username).exists():
            errors.append(f'Username "{username}" already exists.')
        if email and Examiner.objects.filter(email=email).exists():
            errors.append(f'Email "{email}" already exists.')

        if errors:
            for err in errors:
                messages.error(request, err)
            return render(request, 'creator/coordinators/form.html', {
                'default_password': default_password,
                'departments': departments,
                'post': request.POST,
            })

        dept = get_object_or_404(Department, pk=dept_id)

        # Enforce single Head per department
        if position == 'head' and Examiner.objects.filter(
            department=dept, coordinator_position='head',
            role='coordinator', is_deleted=False
        ).exists():
            messages.error(
                request,
                f'Department "{dept.name}" already has a Head coordinator. '
                'Choose a different department or use the RTA position.'
            )
            return render(request, 'creator/coordinators/form.html', {
                'default_password': default_password,
                'departments': departments,
                'post': request.POST,
            })

        coordinator = Examiner(
            username=username,
            email=email or '',
            full_name=full_name,
            role='coordinator',
            department=dept,
            coordinator_position=position,
            is_active=True,
        )
        coordinator.save()

        messages.success(request, f'Coordinator "{full_name}" created successfully.')
        return redirect('creator:coordinator_list')

    return render(request, 'creator/coordinators/form.html', {
        'default_password': default_password,
        'departments': departments,
        'post': {},
    })


@login_required
def coordinator_edit(request, coordinator_id):
    """Edit coordinator details including department and position. Admin/superuser only."""
    if not (request.user.is_superuser or getattr(request.user, 'role', None) == 'admin'):
        messages.error(request, 'You do not have permission to edit coordinators.')
        return redirect('creator:coordinator_list')

    coordinator = get_object_or_404(Examiner, pk=coordinator_id, role='coordinator')
    departments = Department.objects.order_by('name')

    if request.method == 'POST':
        full_name = request.POST.get('full_name', '').strip()
        email = request.POST.get('email', '').strip().lower()
        is_active = request.POST.get('is_active') == 'on'
        dept_id = request.POST.get('department', '').strip()
        position = request.POST.get('coordinator_position', '').strip()

        errors = []
        if not full_name:
            errors.append('Full name is required.')
        if not dept_id:
            errors.append('Department is required.')
        if position not in ('head', 'rta', 'organizer'):
            errors.append('Position must be Head, RTA, or Organizer.')
        if email and email != coordinator.email and Examiner.objects.filter(
            email=email
        ).exclude(id=coordinator_id).exists():
            errors.append(f'Email "{email}" is already used by another user.')

        if errors:
            for err in errors:
                messages.error(request, err)
            return render(request, 'creator/coordinators/edit.html', {
                'coordinator': coordinator,
                'departments': departments,
            })

        dept = get_object_or_404(Department, pk=dept_id)

        # Enforce single Head per department (excluding self)
        if position == 'head' and Examiner.objects.filter(
            department=dept, coordinator_position='head',
            role='coordinator', is_deleted=False
        ).exclude(pk=coordinator_id).exists():
            messages.error(
                request,
                f'Department "{dept.name}" already has a Head coordinator.'
            )
            return render(request, 'creator/coordinators/edit.html', {
                'coordinator': coordinator,
                'departments': departments,
            })

        coordinator.full_name = full_name
        coordinator.email = email or ''
        coordinator.is_active = is_active
        coordinator.department = dept
        coordinator.coordinator_position = position
        coordinator.save()

        messages.success(request, f'Coordinator "{full_name}" updated successfully.')
        return redirect('creator:coordinator_list')

    return render(request, 'creator/coordinators/edit.html', {
        'coordinator': coordinator,
        'departments': departments,
    })


@login_required
def coordinator_delete(request, coordinator_id):
    """Permanently delete a coordinator. Superuser only."""
    if not request.user.is_superuser:
        messages.error(request, 'Only superusers can delete coordinators.')
        return redirect('creator:coordinator_list')

    if request.method != 'POST':
        messages.error(request, 'Invalid request method.')
        return redirect('creator:coordinator_list')

    coordinator = get_object_or_404(Examiner, pk=coordinator_id, role='coordinator')
    full_name = coordinator.full_name
    coordinator.delete()
    messages.success(request, f'Coordinator "{full_name}" deleted successfully.')
    return redirect('creator:coordinator_list')


# ── Department management (admin/superuser only) ───────────────────────────────


@login_required
def department_list(request):
    """List all departments."""
    if not (request.user.is_superuser or getattr(request.user, 'role', None) == 'admin'):
        messages.error(request, 'You do not have permission to manage departments.')
        return redirect('creator:dashboard')

    departments = Department.objects.order_by('name')
    for dept in departments:
        dept.head_count      = dept.members.filter(coordinator_position='head',      is_deleted=False).count()
        dept.rta_count       = dept.members.filter(coordinator_position='rta',       is_deleted=False).count()
        dept.organizer_count = dept.members.filter(coordinator_position='organizer', is_deleted=False).count()

    return render(request, 'creator/departments/list.html', {
        'departments': departments,
    })


@login_required
def department_create(request):
    """Create a new department."""
    if not (request.user.is_superuser or getattr(request.user, 'role', None) == 'admin'):
        messages.error(request, 'You do not have permission to add departments.')
        return redirect('creator:dashboard')

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()

        if not name:
            messages.error(request, 'Department name is required.')
            return render(request, 'creator/departments/form.html', {'post': request.POST})

        if Department.objects.filter(name__iexact=name).exists():
            messages.error(request, f'A department named "{name}" already exists.')
            return render(request, 'creator/departments/form.html', {'post': request.POST})

        dept = Department.objects.create(name=name)
        invalidate_departments()
        invalidate_examiner_list()

        # Auto-create a dry-marking user for this department
        slug = name.lower().replace(' ', '_')
        dry_username = f'dry_{slug}'
        if not Examiner.objects.filter(username=dry_username).exists():
            dry_user = Examiner(
                username=dry_username,
                full_name=f'Dry {name}',
                email='',
                role='examiner',
                department=dept,
                is_active=True,
                is_dry_user=True,
            )
            dry_user.save()  # signal sets default password with must_change_password=False

        messages.success(request, f'Department "{dept.name}" created.')
        return redirect('creator:department_list')

    return render(request, 'creator/departments/form.html', {'post': {}})


@login_required
def department_edit(request, department_id):
    """Edit a department."""
    if not (request.user.is_superuser or getattr(request.user, 'role', None) == 'admin'):
        messages.error(request, 'You do not have permission to edit departments.')
        return redirect('creator:department_list')

    dept = get_object_or_404(Department, pk=department_id)

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()

        if not name:
            messages.error(request, 'Department name is required.')
            return render(request, 'creator/departments/edit.html', {'dept': dept})

        if Department.objects.filter(name__iexact=name).exclude(pk=department_id).exists():
            messages.error(request, f'A department named "{name}" already exists.')
            return render(request, 'creator/departments/edit.html', {'dept': dept})

        dept.name = name
        dept.save()
        invalidate_departments()
        invalidate_examiner_list()
        messages.success(request, f'Department "{dept.name}" updated.')
        return redirect('creator:department_list')

    return render(request, 'creator/departments/edit.html', {'dept': dept})


@login_required
def department_delete(request, department_id):
    """Delete a department. Superuser only. Blocked if coordinators are assigned."""
    if not request.user.is_superuser:
        messages.error(request, 'Only superusers can delete departments.')
        return redirect('creator:department_list')

    if request.method != 'POST':
        return redirect('creator:department_list')

    dept = get_object_or_404(Department, pk=department_id)

    # Only block deletion if there are actual coordinators (with head or rta position)
    coordinators = dept.members.filter(
        is_deleted=False,
        coordinator_position__in=['head', 'rta']
    )
    if coordinators.exists():
        messages.error(
            request,
            f'Cannot delete "{dept.name}" — it still has coordinators assigned. '
            'Reassign or remove them first.'
        )
        return redirect('creator:department_list')

    name = dept.name
    dept.delete()
    invalidate_departments()
    invalidate_examiner_list()
    messages.success(request, f'Department "{name}" deleted.')
    return redirect('creator:department_list')
