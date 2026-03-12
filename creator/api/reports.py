"""
Creator API – Reports & export endpoints.
"""
import csv
import re
from datetime import datetime
from io import BytesIO, StringIO

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404

from core.models import (
    ChecklistItem,
    ExamSession,
    Examiner,
    ItemScore,
    Path,
    SessionStudent,
    Station,
    StationScore,
)


def _safe_filename(name: str) -> str:
    """Sanitize a string for safe use in Content-Disposition headers."""
    return re.sub(r'[^\w\-.]', '_', name)


@login_required
def get_session_summary(request, session_id):
    """GET /api/creator/reports/session/<id>/summary - with pagination and search"""
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    
    try:
        session = get_object_or_404(
            ExamSession.objects.select_related('exam__course'), pk=session_id
        )
        exam_weight = float(session.exam.exam_weight or 0)
        pass_threshold = session.exam.course.pass_threshold

        # Get all students and apply search filter
        search_query = request.GET.get('search', '').strip()
        students_qs = SessionStudent.objects.filter(session=session).order_by('full_name')
        
        if search_query:
            students_qs = students_qs.filter(
                full_name__icontains=search_query
            ) | SessionStudent.objects.filter(
                session=session, student_number__icontains=search_query
            ).order_by('full_name')
            students_qs = students_qs.order_by('full_name').distinct()
        
        # Pagination: 20 students per page
        page_num = request.GET.get('page', 1)
        paginator = Paginator(students_qs, per_page=20)
        try:
            page_obj = paginator.page(page_num)
        except (PageNotAnInteger, EmptyPage):
            page_obj = paginator.page(1)
        
        students = list(page_obj.object_list)

        # Unique station numbers across all paths
        paths = list(Path.objects.filter(session=session))
        station_headers = []
        seen_nums = set()
        for p in paths:
            for s in p.stations.filter(active=True, is_deleted=False):
                if s.station_number not in seen_nums:
                    station_headers.append({'number': s.station_number, 'name': s.name})
                    seen_nums.add(s.station_number)
        station_headers.sort(key=lambda x: x['number'])

        # Pre-calculate station info
        station_info_map = {}
        for p in paths:
            for s in p.stations.filter(active=True, is_deleted=False):
                station_info_map[(p.id, s.station_number)] = {
                    'id': s.id,
                    'max_score': s.get_max_score(),
                }

        total_students = len(students)
        completed_students = 0
        total_percentage = 0
        passed_count = 0

        student_data = []
        # P3: Pre-fetch all paths to avoid N+1 per student
        path_ids = set(s.path_id for s in students if s.path_id)
        path_map = {p.id: p for p in Path.objects.filter(pk__in=path_ids)} if path_ids else {}

        for student in students:
            student_scores = {}
            total_score = 0
            max_score = 0

            for header in station_headers:
                s_info = station_info_map.get((student.path_id, header['number']))
                if s_info:
                    st_max = s_info['max_score']
                    max_score += st_max
                    final = StationScore.get_final_score(student.id, s_info['id'])
                    if final:
                        student_scores[header['number']] = final['final_score']
                        total_score += final['final_score']
                    else:
                        student_scores[header['number']] = None
                else:
                    student_scores[header['number']] = None

            percentage = (total_score / max_score * 100) if max_score > 0 else 0

            # A student is "completed" only if ALL their stations have a submitted score
            scored_stations = sum(1 for v in student_scores.values() if v is not None)
            total_stations = len(student_scores)
            is_completed = total_stations > 0 and scored_stations == total_stations

            path = path_map.get(student.path_id) if student.path_id else None

            weighted_score = round(total_score / max_score * exam_weight, 2) if exam_weight and max_score > 0 else None
            threshold_ratio = pass_threshold / 100
            if exam_weight and weighted_score is not None:
                passed = weighted_score >= exam_weight * threshold_ratio
            else:
                passed = percentage >= pass_threshold

            if is_completed:
                completed_students += 1
                total_percentage += percentage
                if passed:
                    passed_count += 1

            student_data.append({
                'id': str(student.id),
                'student_number': student.student_number,
                'full_name': student.full_name,
                'path_name': path.name if path else None,
                'station_scores': student_scores,
                'total_score': round(total_score, 2),
                'max_score': round(max_score, 2),
                'percentage': round(percentage, 2),
                'weighted_score': weighted_score,
                'passed': passed,
            })

        avg_percentage = (total_percentage / completed_students) if completed_students > 0 else 0
        pass_rate = (passed_count / completed_students * 100) if completed_students > 0 else 0

        return JsonResponse({
            'success': True,
            'data': {
                'session_id': str(session_id),
                'session_name': session.name,
                'exam_weight': exam_weight,
                'total_students': page_obj.paginator.count,
                'completed_students': completed_students,
                'average_percentage': round(avg_percentage, 2),
                'pass_rate': round(pass_rate, 2),
                'station_headers': station_headers,
                'students': sorted(student_data, key=lambda x: x['full_name']),
                'pagination': {
                    'current_page': page_obj.number,
                    'total_pages': page_obj.paginator.num_pages,
                    'total_count': page_obj.paginator.count,
                    'per_page': 20,
                    'has_previous': page_obj.has_previous(),
                    'has_next': page_obj.has_next(),
                    'previous_page_number': page_obj.previous_page_number() if page_obj.has_previous() else None,
                    'next_page_number': page_obj.next_page_number() if page_obj.has_next() else None,
                },
                'search_query': search_query,
            },
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


# ── helpers ─────────────────────────────────────────────────────────────────
def _build_station_info(session_id):
    """Return (paths, station_headers, station_info_map) for a session."""
    paths = list(Path.objects.filter(session_id=session_id))
    station_headers = []
    seen = set()
    station_info_map = {}

    for p in paths:
        for s in p.stations.filter(active=True, is_deleted=False):
            station_info_map[(p.id, s.station_number)] = {
                'id': s.id,
                'max_score': s.get_max_score(),
            }
            if s.station_number not in seen:
                station_headers.append({'number': s.station_number, 'name': s.name})
                seen.add(s.station_number)

    station_headers.sort(key=lambda x: x['number'])
    return paths, station_headers, station_info_map


def _student_rows(students, station_headers, station_info_map, pass_threshold=70):
    """Yield (row_list, total_score, max_score, percentage, pass_fail) for each student."""
    # P3: Pre-fetch paths
    path_ids = set(s.path_id for s in students if s.path_id)
    path_map = {p.id: p for p in Path.objects.filter(pk__in=path_ids)} if path_ids else {}

    for student in students:
        row = [student.student_number, student.full_name]
        path = path_map.get(student.path_id) if student.path_id else None
        row.append(path.name if path else '')

        total_score = 0
        max_score = 0

        for header in station_headers:
            s_info = station_info_map.get((student.path_id, header['number']))
            score_val = ''
            if s_info:
                st_max = s_info['max_score']
                max_score += st_max
                final = StationScore.get_final_score(student.id, s_info['id'])
                if final:
                    score_val = round(final['final_score'], 2)
                    total_score += final['final_score']
                else:
                    score_val = ''
            row.append(score_val)

        percentage = (total_score / max_score * 100) if max_score > 0 else 0
        pass_fail = 'PASS' if percentage >= pass_threshold else 'FAIL'
        row.extend([round(total_score, 2), round(max_score, 2), pass_fail])
        yield row, total_score, max_score, percentage, pass_fail


# ── CSV exports ─────────────────────────────────────────────────────────────

@login_required
def export_students_csv(request, session_id):
    """GET /api/creator/reports/session/<id>/students/csv"""
    session = get_object_or_404(ExamSession.objects.select_related('exam__course'), pk=session_id)
    course_threshold = session.exam.course.pass_threshold
    students = SessionStudent.objects.filter(session=session).order_by('full_name')
    _, station_headers, station_info_map = _build_station_info(session_id)

    output = StringIO()
    writer = csv.writer(output)

    headers = ['Student Number', 'Full Name', 'Path']
    headers.extend([f"St.{h['number']}: {h['name']}" for h in station_headers])
    headers.extend(['Total Score', 'Max Score', 'Pass/Fail'])
    writer.writerow(headers)

    for row, *_ in _student_rows(students, station_headers, station_info_map, pass_threshold=course_threshold):
        writer.writerow(row)

    filename = _safe_filename(f"{session.name}_students_{session_id}.csv")
    response = HttpResponse(output.getvalue(), content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def export_students_xlsx(request, session_id):
    """GET /api/creator/reports/session/<id>/students/xlsx"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    session = get_object_or_404(
        ExamSession.objects.select_related('exam__course'), pk=session_id
    )
    exam_weight = float(session.exam.exam_weight or 0)
    course_threshold = session.exam.course.pass_threshold
    students = list(SessionStudent.objects.filter(session=session).order_by('full_name'))
    _, station_headers, station_info_map = _build_station_info(session_id)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Student Results'

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_alignment = Alignment(horizontal='center', vertical='center')

    base_headers = ['Student Number', 'Full Name', 'Path']
    station_cols = [f"St.{h['number']}: {h['name']}" for h in station_headers]
    weighted_header = [f'Weighted Score (/{exam_weight})'] if exam_weight else []
    end_headers = ['Total Score', 'Max Score'] + weighted_header + ['Pass/Fail', 'Comments']
    all_headers = base_headers + station_cols + end_headers

    # Title rows
    last_col_letter = chr(64 + min(len(all_headers), 26))
    ws.merge_cells(f'A1:{last_col_letter}1')
    ws['A1'] = f"{session.name} - Student Results"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells(f'A2:{last_col_letter}2')
    ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A2'].alignment = Alignment(horizontal='center')

    weighted_fill = PatternFill(start_color='1A3A5C', end_color='1A3A5C', fill_type='solid')
    for col_num, header in enumerate(all_headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.fill = weighted_fill if exam_weight and header.startswith('Weighted') else header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    row_num = 5
    # P3: Pre-fetch paths and stations to avoid N+1
    path_ids = set(s.path_id for s in students if s.path_id)
    path_map = {p.id: p for p in Path.objects.filter(pk__in=path_ids)} if path_ids else {}
    station_id_set = set()
    for info in station_info_map.values():
        station_id_set.add(info['id'])
    station_obj_map = {s.id: s for s in Station.objects.filter(pk__in=station_id_set)} if station_id_set else {}

    for student in students:
        # Get all scores for this student
        # Only count submitted scores, ignore abandoned/in-progress evaluations
        scores_qs = StationScore.objects.filter(
            session_student=student,
            status='submitted'
        ).exclude(
            total_score__isnull=True
        ).select_related('examiner')
        
        # Build row with station scores
        row = [student.student_number, student.full_name]
        path = path_map.get(student.path_id) if student.path_id else None
        row.append(path.name if path else '')

        total_score = 0
        max_score = 0

        for header in station_headers:
            s_info = station_info_map.get((student.path_id, header['number']))
            score_val = ''
            if s_info:
                st_max = s_info['max_score']
                max_score += st_max
                final = StationScore.get_final_score(student.id, s_info['id'])
                if final:
                    score_val = round(final['final_score'], 2)
                    total_score += final['final_score']
                else:
                    score_val = ''
            row.append(score_val)

        percentage = (total_score / max_score * 100) if max_score > 0 else 0
        weighted_score = round(total_score / max_score * exam_weight, 2) if exam_weight and max_score > 0 else None
        # Pass if weighted score >= threshold% of exam_weight, or raw percentage >= threshold%
        threshold_ratio = course_threshold / 100
        if exam_weight and weighted_score is not None:
            passed = weighted_score >= exam_weight * threshold_ratio
        else:
            passed = percentage >= course_threshold
        pass_fail = 'PASS' if passed else 'FAIL'
        
        # Build comments with examiner names
        comments_list = []
        for score in scores_qs:
            if score.comments and score.comments.strip():
                examiner_name = score.examiner.full_name if score.examiner else "Unknown Examiner"
                station_obj = station_obj_map.get(score.station_id)
                station_name = station_obj.name if station_obj else "Unknown Station"
                comments_list.append(f"{examiner_name} ({station_name}):\n{score.comments}")
        
        comments_text = "\n---\n".join(comments_list) if comments_list else ""
        
        weighted_cols = [weighted_score] if exam_weight else []
        row.extend([round(total_score, 2), round(max_score, 2)] + weighted_cols + [pass_fail, comments_text])
        
        # Write row to worksheet
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # Color-code pass/fail
        pf_col_idx = len(all_headers) - 1  # Pass/Fail is second-to-last (before Comments)
        pf_cell = ws.cell(row=row_num, column=pf_col_idx)
        if pass_fail == 'PASS':
            pf_cell.font = Font(color='006100', bold=True)
            pf_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        else:
            pf_cell.font = Font(color='9C0006', bold=True)
            pf_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

        # Highlight weighted score column
        if exam_weight:
            ws_col_idx = pf_col_idx - 1  # Weighted Score is just before Pass/Fail
            ws_cell = ws.cell(row=row_num, column=ws_col_idx)
            ws_cell.font = Font(bold=True, color='1A1A2E')
            ws_cell.fill = PatternFill(start_color='E8F4FD', end_color='E8F4FD', fill_type='solid')
        
        row_num += 1

    # Column widths: fixed per column type
    num_base = len(base_headers)          # Student #, Full Name, Path
    num_station = len(station_cols)        # One per station
    # base columns
    ws.column_dimensions['A'].width = 16  # Student Number
    ws.column_dimensions['B'].width = 28  # Full Name
    ws.column_dimensions['C'].width = 14  # Path
    # station score columns — narrow numeric
    for i in range(num_station):
        ws.column_dimensions[get_column_letter(num_base + 1 + i)].width = 10
    # end columns
    end_start = num_base + num_station + 1
    ws.column_dimensions[get_column_letter(end_start)].width = 13      # Total Score
    ws.column_dimensions[get_column_letter(end_start + 1)].width = 13  # Max Score
    if exam_weight:
        ws.column_dimensions[get_column_letter(end_start + 2)].width = 18  # Weighted Score
        ws.column_dimensions[get_column_letter(end_start + 3)].width = 10  # Pass/Fail
        ws.column_dimensions[get_column_letter(end_start + 4)].width = 50  # Comments
    else:
        ws.column_dimensions[get_column_letter(end_start + 2)].width = 10  # Pass/Fail
        ws.column_dimensions[get_column_letter(end_start + 3)].width = 50  # Comments

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = _safe_filename(f"{session.name}_students_{session_id}.xlsx")
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def export_stations_csv(request, session_id):
    """GET /api/creator/reports/session/<id>/stations/csv"""
    session = get_object_or_404(ExamSession, pk=session_id)
    # P5: Prefetch paths and stations in one query
    stations = Station.objects.filter(
        path__session_id=session_id, active=True
    ).select_related('path')

    output = StringIO()
    writer = csv.writer(output)
    writer.writerow([
        'Station Name', 'Path', 'Max Score', 'Avg Score', 'Avg %',
        'Min Score', 'Max Achieved', 'Students Marked',
    ])

    for station in stations:
        scores = StationScore.objects.filter(station=station)
        if scores.exists():
            total_scores = [s.total_score or 0 for s in scores]
            max_scores = [s.max_score or 0 for s in scores]
            avg_score = sum(total_scores) / len(total_scores)
            avg_max = sum(max_scores) / len(max_scores) if max_scores else 0
            avg_pct = (avg_score / avg_max * 100) if avg_max > 0 else 0
            writer.writerow([
                station.name,
                station.path.name if station.path else '',
                round(avg_max, 2),
                round(avg_score, 2),
                round(avg_pct, 2),
                round(min(total_scores), 2),
                round(max(total_scores), 2),
                len(total_scores),
            ])

    filename = _safe_filename(f"{session.name}_stations_{session_id}.csv")
    response = HttpResponse(output.getvalue(), content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def export_raw_csv(request, session_id):
    """GET /api/creator/reports/session/<id>/raw/csv"""
    session = get_object_or_404(ExamSession, pk=session_id)

    # P1: Eliminate N+1 queries — prefetch everything in a single query chain
    item_scores = ItemScore.objects.filter(
        station_score__session_student__session=session
    ).select_related(
        'station_score__session_student',
        'station_score__examiner',
        'station_score__station',
        'checklist_item',
    ).order_by(
        'station_score__session_student__student_number',
        'station_score__station__station_number',
        'checklist_item__item_number',
    )

    # Pre-fetch path names for all students in one query
    path_map = {}
    student_path_ids = set(
        SessionStudent.objects.filter(session=session)
        .exclude(path_id__isnull=True)
        .values_list('path_id', flat=True)
    )
    if student_path_ids:
        for p in Path.objects.filter(pk__in=student_path_ids):
            path_map[p.id] = p.name

    output = StringIO()
    writer = csv.writer(output)
    writer.writerow([
        'Student Number', 'Student Name', 'Path', 'Station',
        'Item', 'Score', 'Max Score', 'Examiner', 'Timestamp',
    ])

    for iscore in item_scores:
        ss = iscore.station_score
        student = ss.session_student
        writer.writerow([
            student.student_number,
            student.full_name,
            path_map.get(student.path_id, ''),
            ss.station.name if ss.station else '',
            iscore.checklist_item.description if iscore.checklist_item else '',
            iscore.score or 0,
            iscore.max_points or 0,       # S8: was iscore.max_score (wrong field)
            ss.examiner.full_name if ss.examiner else '',
            iscore.marked_at or '',        # S8: was iscore.scored_at (wrong field)
        ])

    filename = _safe_filename(f"{session.name}_raw_{session_id}.csv")
    response = HttpResponse(output.getvalue(), content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
