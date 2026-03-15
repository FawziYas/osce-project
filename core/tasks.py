"""
Celery tasks for the OSCE project.

Tasks:
  core.write_audit_log         – async audit log write (original)
  core.cleanup_old_audit_logs  – periodic: delete audit entries > 90 days
  core.compute_dashboard_stats – periodic: pre-compute homepage stats
  core.check_session_readiness – one-off: validate session before activation
  core.bulk_import_examiners   – async: process uploaded XLSX in background
  core.generate_pdf_report     – async: generate session PDF in background
"""
import logging
import traceback

from celery import shared_task

logger = logging.getLogger('osce.audit')


@shared_task(
    name='core.write_audit_log',
    bind=True,
    max_retries=3,
    default_retry_delay=5,
    acks_late=True,
    ignore_result=True,
)
def write_audit_log(self, payload):
    """
    Write an audit log entry to the database.

    Called from AuditLogService.log() when Celery is available.
    The payload is a plain dict (JSON-serialisable) to avoid
    pickling Django model instances.
    """
    try:
        from core.models.audit import AuditLog

        AuditLog.objects.create(
            user_id=payload.get('user_id'),
            username=payload.get('username', ''),
            user_role=payload.get('user_role', ''),
            department_id=payload.get('department_id'),
            action=payload.get('action', ''),
            status=payload.get('status', 'SUCCESS'),
            resource_type=payload.get('resource_type', ''),
            resource_id=payload.get('resource_id', ''),
            resource_label=payload.get('resource_label', ''),
            old_value=payload.get('old_value'),
            new_value=payload.get('new_value'),
            description=payload.get('description', ''),
            ip_address=payload.get('ip_address'),
            user_agent=payload.get('user_agent') or '',
            request_method=payload.get('request_method') or '',
            request_path=payload.get('request_path') or '',
            extra_data=payload.get('extra_data'),
        )

    except Exception as exc:
        logger.error(
            'Celery write_audit_log failed (attempt %d): %s',
            self.request.retries + 1,
            traceback.format_exc(),
        )
        raise self.retry(exc=exc)

# ══════════════════════════════════════════════════════════════════════════════
# 2. Periodic: cleanup old audit logs (runs nightly via Celery Beat)
# ══════════════════════════════════════════════════════════════════════════════

@shared_task(name='core.cleanup_old_audit_logs', ignore_result=True)
def cleanup_old_audit_logs(days=365):
    """Delete AuditLog entries older than `days` days (default: 1 year). Scheduled nightly."""
    try:
        from django.utils import timezone
        from datetime import timedelta
        from core.models.audit import AuditLog

        cutoff = timezone.now() - timedelta(days=days)
        deleted_count, _ = AuditLog.objects.filter(timestamp__lt=cutoff).delete()
        logger.info('Audit log cleanup: deleted %d entries older than %d days', deleted_count, days)
        return {'deleted': deleted_count}
    except Exception:
        logger.error('cleanup_old_audit_logs failed: %s', traceback.format_exc())


# ══════════════════════════════════════════════════════════════════════════════
# 3. Periodic: pre-compute dashboard statistics (runs every 5 min via Beat)
# ══════════════════════════════════════════════════════════════════════════════

@shared_task(name='core.compute_dashboard_stats', ignore_result=True)
def compute_dashboard_stats():
    """Pre-compute and cache global dashboard statistics every 5 minutes."""
    try:
        from datetime import date
        from django.core.cache import cache
        from core.models import Examiner, ExamSession, ExaminerAssignment, SessionStudent
        from core.utils.cache_utils import DASHBOARD_STATS_KEY, DASHBOARD_STATS_TTL

        today = date.today()
        stats = {
            'total_examiners': Examiner.objects.filter(role='examiner', is_deleted=False).count(),
            'active_examiners': Examiner.objects.filter(role='examiner', is_active=True, is_deleted=False).count(),
            'sessions_today': ExamSession.objects.filter(session_date=today).count(),
            'sessions_in_progress': ExamSession.objects.filter(status='in_progress').count(),
            'assignments_today': ExaminerAssignment.objects.filter(
                session__session_date=today,
            ).values('examiner_id').distinct().count(),
            'students_today': SessionStudent.objects.filter(session__session_date=today).count(),
        }
        cache.set(DASHBOARD_STATS_KEY, stats, DASHBOARD_STATS_TTL)
        logger.info('Dashboard stats refreshed: %s', stats)
        return stats
    except Exception:
        logger.error('compute_dashboard_stats failed: %s', traceback.format_exc())


# ══════════════════════════════════════════════════════════════════════════════
# 4. One-off: session readiness check (called before activation)
# ══════════════════════════════════════════════════════════════════════════════

@shared_task(name='core.check_session_readiness', bind=True, max_retries=2, default_retry_delay=3)
def check_session_readiness(self, session_id):
    """
    Validate a session is ready for activation. Returns:
      ready: bool, errors: list, warnings: list
    """
    try:
        from core.models import ExamSession, Path, ExaminerAssignment

        session = ExamSession.objects.select_related('exam').get(pk=session_id)
        paths = list(Path.objects.filter(session=session, is_deleted=False))
        assignments = ExaminerAssignment.objects.filter(session=session)
        errors = []
        warnings = []

        if not paths:
            errors.append('No rotation paths have been created.')

        total_students = session.students.count()
        if total_students == 0:
            errors.append('No students have been added to this session.')

        unassigned_students = session.students.filter(path__isnull=True).count()
        if unassigned_students > 0:
            warnings.append(f'{unassigned_students} student(s) not assigned to any path.')

        assigned_station_ids = set(assignments.values_list('station_id', flat=True))
        unassigned_stations = []
        for path in paths:
            if path.stations.filter(active=True, is_deleted=False).count() == 0:
                errors.append(f'Path {path.name} has no active stations.')
            for station in path.stations.filter(active=True, is_deleted=False):
                if station.id not in assigned_station_ids:
                    unassigned_stations.append(f'{path.name}/{station.name}')

        if unassigned_stations:
            preview = ', '.join(unassigned_stations[:5])
            extra = f' ...(+{len(unassigned_stations) - 5} more)' if len(unassigned_stations) > 5 else ''
            warnings.append(f'{len(unassigned_stations)} station(s) without examiner: {preview}{extra}')

        result = {
            'session_id': str(session_id),
            'session_name': session.name,
            'ready': len(errors) == 0,
            'errors': errors,
            'warnings': warnings,
            'total_students': total_students,
            'total_paths': len(paths),
            'total_assignments': assignments.count(),
        }
        logger.info('Session readiness check %s: ready=%s', session_id, result['ready'])
        return result
    except Exception as exc:
        logger.error('check_session_readiness failed: %s', traceback.format_exc())
        raise self.retry(exc=exc)


# ══════════════════════════════════════════════════════════════════════════════
# 5. Async: bulk XLSX examiner import
# ══════════════════════════════════════════════════════════════════════════════

@shared_task(
    name='core.bulk_import_examiners',
    bind=True,
    max_retries=1,
    time_limit=120,
    soft_time_limit=100,
)
def bulk_import_examiners(self, file_path, requested_by_id=None):
    """
    Process a saved XLSX file and import examiners asynchronously.
    Stores progress/result in cache under 'osce:bulk_import:<task_id>'.
    """
    import os
    import openpyxl
    from django.core.cache import cache

    task_id = self.request.id
    result_key = f'osce:bulk_import:{task_id}'

    def _save(payload):
        cache.set(result_key, payload, 60 * 30)

    _save({'status': 'running', 'progress': 0})

    try:
        from core.models import Examiner, Department
        from core.utils.cache_utils import invalidate_examiner_list

        valid_departments = set(Department.objects.values_list('name', flat=True))
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        def _sanitize_cell(val):
            """Strip formula-triggering prefixes from imported cell values."""
            if not isinstance(val, str):
                return val
            if val and val[0] in ('=', '+', '@', '\t', '\r'):
                return val.lstrip('=+@\t\r')
            return val

        headers = [str(c.value).strip().lower() for c in ws[1] if c.value]
        idx = {h: i for i, h in enumerate(headers)}
        rows = list(ws.iter_rows(min_row=3, values_only=True))
        total = max(len(rows), 1)
        success_count = 0
        errors_list = []

        for offset, row in enumerate(rows):
            row_num = offset + 3
            if not any(row):
                continue
            try:
                get = lambda col, default='': _sanitize_cell(str(row[idx[col]]).strip()) if col in idx and idx[col] < len(row) and row[idx[col]] else default
                username = get('username').lower()
                full_name = get('full_name')
                email = get('email').lower()
                title = get('title')
                department = get('department')

                if not username or not full_name:
                    errors_list.append(f'Row {row_num}: Missing username or full_name')
                    continue
                if department and department not in valid_departments:
                    errors_list.append(f"Row {row_num}: Department '{department}' not found")
                    continue
                if Examiner.objects.filter(username=username).exists():
                    errors_list.append(f"Row {row_num}: Username '{username}' already exists")
                    continue
                if email and Examiner.objects.filter(email=email).exists():
                    errors_list.append(f"Row {row_num}: Email '{email}' already exists")
                    continue
                Examiner.objects.create(
                    username=username, email=email, full_name=full_name,
                    title=title, department=department, is_active=True,
                )
                success_count += 1
            except Exception as e:
                errors_list.append(f'Row {row_num}: {e}')

            if offset % 10 == 0:
                _save({'status': 'running', 'progress': int((offset / total) * 100)})

        invalidate_examiner_list()
        final = {
            'status': 'done', 'progress': 100,
            'success_count': success_count,
            'error_count': len(errors_list),
            'errors': errors_list[:20],
        }
        _save(final)
        logger.info('Bulk import done: %d added, %d errors', success_count, len(errors_list))
        return final

    except Exception as exc:
        _save({'status': 'error', 'message': str(exc)})
        logger.error('bulk_import_examiners failed: %s', traceback.format_exc())
        raise
    finally:
        try:
            if os.path.exists(file_path):
                os.remove(file_path)
        except Exception:
            pass


# ══════════════════════════════════════════════════════════════════════════════
# 6. Async: PDF report generation
# ══════════════════════════════════════════════════════════════════════════════

@shared_task(
    name='core.generate_pdf_report',
    bind=True,
    max_retries=2,
    default_retry_delay=5,
    time_limit=180,
    soft_time_limit=150,
)
def generate_pdf_report(self, session_id):
    """
    Generate student-paths PDF for a session in the background.
    Stores base64 PDF in cache key 'osce:pdf_status:<session_id>:<task_id>'.
    Frontend polls GET /creator/sessions/<id>/pdf-status/?task_id=<uuid>
    """
    import base64
    from django.core.cache import cache

    task_id = self.request.id
    status_key = f'osce:pdf_status:{session_id}:{task_id}'
    cache.set(status_key, {'status': 'running'}, 60 * 10)

    try:
        from core.models import ExamSession
        from creator.views.sessions import _build_student_paths_pdf

        session = ExamSession.objects.get(pk=session_id)
        pdf_bytes = _build_student_paths_pdf(session)
        pdf_b64 = base64.b64encode(pdf_bytes).decode('utf-8')
        cache.set(status_key, {
            'status': 'done',
            'pdf_b64': pdf_b64,
            'filename': f'student_paths_{session_id}.pdf',
        }, 60 * 15)
        logger.info('PDF report generated for session %s', session_id)
        return {'status': 'done'}

    except Exception as exc:
        cache.set(status_key, {'status': 'error', 'message': str(exc)}, 60 * 5)
        logger.error('generate_pdf_report failed for %s: %s', session_id, traceback.format_exc())
        raise self.retry(exc=exc)